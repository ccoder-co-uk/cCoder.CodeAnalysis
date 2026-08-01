import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(r"C:\Data\Github\cCoder")
OUTPUT = Path(__file__).resolve().parent / "cCoder RFC Rule Candidate Audit.xlsx"
SKIP_PARTS = {"_worktrees", "_temp", "obj", "bin", "decompiled"}
CRUD_NAMES = {"Get", "GetAll", "Post", "Put", "Patch", "Delete"}


def included(path):
    return not any(part in SKIP_PARTS for part in path.parts)


def extract_block(text, start):
    brace = text.find("{", start)
    arrow = text.find("=>", start)
    semi = text.find(";", start)
    if arrow >= 0 and (brace < 0 or arrow < brace):
        end = semi if semi >= 0 else min(len(text), arrow + 500)
        return text[start:end + 1]
    if brace < 0:
        return text[start:min(len(text), start + 1000)]
    depth = 0
    for index in range(brace, len(text)):
        if text[index] == "{":
            depth += 1
        elif text[index] == "}":
            depth -= 1
            if depth == 0:
                return text[start:index + 1]
    return text[start:]


def repo_name(path):
    try:
        return path.relative_to(ROOT).parts[0]
    except ValueError:
        return ""


def classify_result(body):
    patterns = [
        ("201 Created", r"\bCreated(?:AtAction|AtRoute)?\s*\("),
        ("204 No Content", r"\bNoContent\s*\("),
        ("OData Updated", r"\bUpdated\s*\("),
        ("200 OK", r"\bOk\s*\("),
        ("404 Not Found", r"\bNotFound\s*\("),
        ("400 Bad Request", r"\bBadRequest(?:Result)?\b"),
    ]
    found = [label for label, pattern in patterns if re.search(pattern, body)]
    return ", ".join(found) if found else "Not statically recognized"


def find_tests(repo, controller, action):
    matches = []
    test_roots = [p for p in repo.rglob("*.cs") if included(p) and "test" in str(p).lower()]
    controller_stem = controller.removesuffix("Controller") + "ControllerTests"
    for path in test_roots:
        name = path.name
        if controller_stem.lower() not in name.lower():
            continue
        if "." + action.lower() not in name.lower() and action.lower() not in path.read_text(encoding="utf-8-sig", errors="ignore").lower():
            continue
        matches.append(path)
    return matches


def test_assertions(paths):
    fragments = []
    for path in paths:
        text = path.read_text(encoding="utf-8-sig", errors="ignore")
        for pattern in [
            r"HttpStatusCode\.\w+", r"StatusCodes\.Status\d+\w+",
            r"StatusCode[^;\n]{0,100}", r"BeOfType<\w+Result>",
            r"Be\(\(int\)HttpStatusCode\.\w+\)",
        ]:
            fragments.extend(re.findall(pattern, text, flags=re.IGNORECASE))
    return "; ".join(dict.fromkeys(x.strip() for x in fragments))


def add_table(ws, name):
    if ws.max_row < 2:
        return
    table = Table(displayName=name, ref=f"A1:{ws.cell(ws.max_row, ws.max_column).coordinate}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    ws.freeze_panes = "A2"


def format_sheet(ws, widths):
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def main():
    controller_rows = []
    method_rows = []
    test_cache = {}
    controller_pattern = re.compile(
        r"\b(?:class|record\s+class)\s+(?P<name>\w+Controller)\b[\s\S]{0,500}?\:\s*(?:[\w.]+\s*,\s*)*ODataController\b")
    method_pattern = re.compile(
        r"(?P<attrs>(?:\s*\[[^\]]+\]\s*)*)\s*public\s+(?:virtual\s+|override\s+|sealed\s+|async\s+)*"
        r"(?P<return>[\w<>,.?\[\]\s]+?)\s+(?P<name>GetAll|Get|Post|Put|Patch|Delete)\s*\((?P<params>[^)]*)\)",
        re.MULTILINE)

    files = [p for p in ROOT.rglob("*.cs") if included(p)]
    for path in files:
        text = path.read_text(encoding="utf-8-sig", errors="ignore")
        class_match = controller_pattern.search(text)
        if not class_match:
            continue
        controller = class_match.group("name")
        repo = ROOT / repo_name(path)
        methods = []
        for match in method_pattern.finditer(text):
            name = match.group("name")
            body = extract_block(text, match.start())
            attrs = match.group("attrs").strip().replace("\n", " ")
            params = " ".join(match.group("params").split())
            entity_match = re.search(r"\[FromBody\]\s*([\w.<>]+)", params)
            entity = entity_match.group(1) if entity_match else ""
            standard_crud = name in CRUD_NAMES and (name != "Post" or bool(entity))
            tests_key = (str(repo), controller, name)
            if tests_key not in test_cache:
                test_cache[tests_key] = find_tests(repo, controller, name)
            tests = test_cache[tests_key]
            method_rows.append([
                repo.name, controller, name, entity, attrs, params,
                classify_result(body), "Yes" if standard_crud else "Review",
                "\n".join(str(p.relative_to(repo)) for p in tests),
                test_assertions(tests), str(path.relative_to(repo)),
            ])
            methods.append(name)
        controller_rows.append([
            repo.name, controller, str(path.relative_to(repo)),
            ", ".join(methods), len(methods),
        ])

    crud_posts = [row for row in method_rows if row[2] == "Post" and row[3]]
    post_results = Counter(row[6] for row in crud_posts)
    crud_counts = Counter(row[2] for row in method_rows if row[7] == "Yes")
    rule_rows = [
        ["RFC0001", "OData CRUD Post(T)", "Return 201 Created with created representation", "OData 4.01 §9.1.2; RFC 9110 §15.3.2", len(crud_posts), sum(1 for r in crud_posts if "201 Created" in r[6]), sum(1 for r in crud_posts if "200 OK" in r[6]), "High", "Approved"],
        ["RFC0002", "OData CRUD Delete(key)", "Return 204 when deletion succeeds without a representation", "RFC 9110 §15.3.5; adopted cCoder policy", crud_counts.get("Delete", 0), sum(1 for r in method_rows if r[2] == "Delete" and "204 No Content" in r[6]), sum(1 for r in method_rows if r[2] == "Delete" and "200 OK" in r[6]), "High", "Approved"],
        ["RFC0003", "OData CRUD Get/GetAll", "Return 200 with a representation for successful retrieval", "RFC 9110 §15.3.1; OData 4.01 §9.1.1", crud_counts.get("Get", 0) + crud_counts.get("GetAll", 0), sum(1 for r in method_rows if r[2].startswith("Get") and "200 OK" in r[6]), 0, "High", "Approved"],
        ["RFC0004", "OData CRUD Put/Patch", "Return the updated representation with 200 OK; OData Updated(entity) is accepted", "RFC 9110 §§9.3.4,15.3.1; OData semantics", crud_counts.get("Put", 0) + crud_counts.get("Patch", 0), sum(1 for r in method_rows if r[2] in {"Put", "Patch"} and "OData Updated" in r[6]), sum(1 for r in method_rows if r[2] in {"Put", "Patch"} and "200 OK" in r[6]), "High", "Approved"],
        ["TEST0001", "Public OData CRUD action", "Matching acceptance-test contract scenario must exist", "cCoder quality gate, not directly an RFC requirement", sum(1 for r in method_rows if r[7] == "Yes"), sum(1 for r in method_rows if r[7] == "Yes" and r[8]), sum(1 for r in method_rows if r[7] == "Yes" and not r[8]), "Medium", "Best enforced in test compilation/meta-test, not production compilation"],
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Rule Candidates"
    ws.append(["Candidate ID", "Inferred Shape", "Proposed Requirement", "Standards Basis", "Applicable Methods", "Observed Target Pattern", "Observed Legacy Pattern", "Inference Confidence", "Decision Status"])
    for row in rule_rows:
        ws.append(row)
    add_table(ws, "RuleCandidates")
    format_sheet(ws, {"A": 28, "B": 30, "C": 55, "D": 48, "E": 18, "F": 24, "G": 24, "H": 20, "I": 45})

    ws = wb.create_sheet("CRUD Methods")
    ws.append(["Repository", "Controller", "Action", "Body Entity Type", "Attributes", "Parameters", "Observed Results", "Standard CRUD Candidate", "Matching Test Files", "Observed Status Assertions", "Controller File"])
    for row in sorted(method_rows):
        ws.append(row)
    add_table(ws, "CrudMethods")
    format_sheet(ws, {"A": 28, "B": 34, "C": 14, "D": 28, "E": 48, "F": 55, "G": 32, "H": 22, "I": 60, "J": 55, "K": 70})

    ws = wb.create_sheet("Controllers")
    ws.append(["Repository", "Controller", "File", "Discovered CRUD Actions", "Action Count"])
    for row in sorted(controller_rows):
        ws.append(row)
    add_table(ws, "Controllers")
    format_sheet(ws, {"A": 30, "B": 36, "C": 75, "D": 55, "E": 15})

    ws = wb.create_sheet("Repository Summary")
    ws.append(["Repository", "OData Controllers", "CRUD Methods", "Entity Posts", "Posts Returning 201", "Posts Returning 200", "Deletes Returning 204", "Deletes Returning 200", "Methods With Likely Test Match", "Methods Without Likely Test Match"])
    repositories = sorted({row[0] for row in controller_rows})
    for repository in repositories:
        repository_controllers = [r for r in controller_rows if r[0] == repository]
        repository_methods = [r for r in method_rows if r[0] == repository and r[7] == "Yes"]
        posts = [r for r in repository_methods if r[2] == "Post" and r[3]]
        deletes = [r for r in repository_methods if r[2] == "Delete"]
        ws.append([
            repository, len(repository_controllers), len(repository_methods), len(posts),
            sum(1 for r in posts if "201 Created" in r[6]),
            sum(1 for r in posts if "200 OK" in r[6]),
            sum(1 for r in deletes if "204 No Content" in r[6]),
            sum(1 for r in deletes if "200 OK" in r[6]),
            sum(1 for r in repository_methods if r[8]),
            sum(1 for r in repository_methods if not r[8]),
        ])
    add_table(ws, "RepositorySummary")
    format_sheet(ws, {"A": 34, "B": 20, "C": 18, "D": 16, "E": 22, "F": 22, "G": 24, "H": 24, "I": 28, "J": 30})

    ws = wb.create_sheet("Read Me")
    for row in [
        ["Item", "Value"],
        ["Purpose", "Evidence report for jointly refining proposed RFCxxxx cCoder.CodeAnalysis rules before implementation."],
        ["Generated UTC", datetime.now(timezone.utc).isoformat()],
        ["Scope", str(ROOT) + " direct repository checkouts; worktrees, generated obj/bin, temp and decompiled sources excluded."],
        ["OData controllers", len(controller_rows)],
        ["CRUD/action methods", len(method_rows)],
        ["Entity Post candidates", len(crud_posts)],
        ["Caution", "Static text analysis is intentionally a policy-discovery aid, not the final Roslyn semantic implementation."],
        ["Interpretation", "Observed target pattern means code already uses the standards-aligned result. Observed legacy pattern identifies likely migration work."],
        ["Test matching", "Filename/controller/action convention is used to locate likely acceptance tests; blank or matched does not prove runtime coverage."],
    ]:
        ws.append(row)
    format_sheet(ws, {"A": 28, "B": 120})

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    checked = load_workbook(OUTPUT, read_only=True)
    assert checked["CRUD Methods"].max_row == len(method_rows) + 1
    print(f"Output: {OUTPUT}")
    print(f"Controllers: {len(controller_rows)}; methods: {len(method_rows)}; entity posts: {len(crud_posts)}")
    print(f"Entity Post observed results: {dict(post_results)}")


if __name__ == "__main__":
    main()
